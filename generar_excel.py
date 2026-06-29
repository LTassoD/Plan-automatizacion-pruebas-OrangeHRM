import openpyxl
import os

# Crea la carpeta testData si no existe
os.makedirs("testData", exist_ok=True)

# --- dataEmpleados.xlsx (Case 5: Registro de empleados) ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Empleados"
ws.append(["NroFila", "Nombre", "Apellido", "Usuario", "Contrasena"])
ws.append([1, "Juan", "Perez", "jperez", "Pass123!"])
ws.append([2, "Maria", "Gonzalez", "mgonzalez", "Pass456!"])
ws.append([3, "Carlos", "Lopez", "clopez", "Pass789!"])
wb.save("testData/dataEmpleados.xlsx")

# --- dataFiltros.xlsx (Case 6: Busqueda con filtros) ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Filtros"
ws.append(["NroFila", "Nombre", "Estado", "ResultadoEsperado"])
ws.append([1, "Admin", "Active", "1 resultado"])
ws.append([2, "XXXXXXXXXXX", "Active", "0 resultados"])
ws.append([3, "Admin", "Active", "1 resultado"])
wb.save("testData/dataFiltros.xlsx")

# --- dataEdicion.xlsx (Case 7: Edicion de empleados) ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Edicion"
ws.append(["NroFila", "ID", "NuevoCargo", "NuevoDepartamento"])
ws.append([1, "0001", "QA Engineer", "IT"])
ws.append([2, "0002", "Developer", "Engineering"])
ws.append([3, "0003", "Product Owner", "Management"])
wb.save("testData/dataEdicion.xlsx")

# --- dataEliminacion.xlsx (Case 8: Eliminacion de empleados) ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Eliminacion"
ws.append(["NroFila", "Nombre", "Apellido", "ResultadoEsperado"])
ws.append([1, "Juan", "Perez", "No aparece en busqueda"])
ws.append([2, "Maria", "Gonzalez", "No aparece en busqueda"])
ws.append([3, "Carlos", "Lopez", "No aparece en busqueda"])
wb.save("testData/dataEliminacion.xlsx")

# --- dataPerfil.xlsx (Case 13: Actualizar perfil My Info) ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Perfil"
ws.append(["NroFila", "Campo", "ValorNuevo"])
ws.append([1, "Nickname", "JuanQA"])
ws.append([2, "DriverLicense", "B-12345"])
ws.append([3, "Nationality", "Chilean"])
wb.save("testData/dataPerfil.xlsx")

# --- dataLicencias.xlsx (Case 14: Solicitud de licencias) ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Licencias"
ws.append(["NroFila", "TipoLicencia", "FechaInicio", "FechaFin", "Resultado"])
ws.append([1, "Annual Leave", "2025-07-01", "2025-07-05", "Aprobada"])
ws.append([2, "Sick Leave", "2025-07-10", "2025-07-11", "Aprobada"])
ws.append([3, "Casual Leave", "2025-07-20", "2025-07-20", "Aprobada"])
wb.save("testData/dataLicencias.xlsx")

print("Todos los archivos Excel generados en testData/")
