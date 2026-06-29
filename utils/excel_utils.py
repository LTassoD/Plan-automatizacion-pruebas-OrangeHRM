import openpyxl


class ExcelUtils:
    """Clase de utilidad para leer datos desde archivos Excel (.xlsx).
    Se usa en los steps Data-Driven para obtener datos de prueba.

    Uso:
        ExcelUtils.set_excel_file_sheet("testData/dataEmpleados.xlsx", "Empleados")
        nombre = ExcelUtils.get_cell_data(2, 1)  # fila 2, columna 1
    """

    _wb = None
    _sheet = None

    @classmethod
    def set_excel_file_sheet(cls, filepath, sheet_name):
        """Carga un archivo Excel y selecciona una hoja especifica.

        Args:
            filepath: Ruta al archivo .xlsx.
            sheet_name: Nombre de la hoja a utilizar.

        Returns:
            True si la hoja existe, False en caso contrario.
        """
        cls._wb = openpyxl.load_workbook(filepath)
        if sheet_name in cls._wb.sheetnames:
            cls._sheet = cls._wb[sheet_name]
            return True
        return False

    @classmethod
    def get_cell_data(cls, row, col):
        """Obtiene el valor de una celda especifica (1-indexed).

        Args:
            row: Numero de fila (1 = primera fila, normalmente encabezados).
            col: Numero de columna (1 = columna A).

        Returns:
            Valor de la celda, o None si no hay hoja cargada.
        """
        if cls._sheet:
            return cls._sheet.cell(row=row, column=col).value
        return None
